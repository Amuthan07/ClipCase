#!/usr/bin/env python3
"""
clipcase — Convert screen recording videos into structured test cases.

Usage:
    python clipcase.py <video_file> [options]

Examples:
    python clipcase.py recording.mov
    python clipcase.py recording.mov --fps 2 --provider anthropic
    python clipcase.py recording.mov --fps 1 --provider openai --output my_tests
"""

import argparse
import base64
import csv
import json
import os
import subprocess
import sys
import glob
import time
from pathlib import Path


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SUPPORTED_VIDEO_EXTS = {".mov", ".mp4", ".webm", ".avi", ".mkv"}
SUPPORTED_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
DEFAULT_FPS = 2
MAX_FRAMES_PER_BATCH = 8  # frames sent to LLM per API call
FRAME_SAMPLE_INTERVAL = 10  # sample every Nth frame for analysis


def load_env():
    """Load .env file if it exists."""
    env_path = Path(__file__).parent / ".env"
    if env_path.exists():
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, _, value = line.partition("=")
                    os.environ.setdefault(key.strip(), value.strip().strip("\"'"))


# ---------------------------------------------------------------------------
# Frame Extraction
# ---------------------------------------------------------------------------

def check_ffmpeg():
    """Verify ffmpeg is installed."""
    try:
        result = subprocess.run(
            ["ffmpeg", "-version"], capture_output=True, text=True, timeout=10
        )
        return result.returncode == 0
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def get_video_info(video_path: str) -> dict:
    """Get video metadata using ffprobe."""
    cmd = [
        "ffprobe", "-v", "quiet",
        "-print_format", "json",
        "-show_format", "-show_streams",
        video_path,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    if result.returncode != 0:
        return {}
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        return {}


def resolve_video_path(video_arg: str) -> str:
    """
    Resolve video file path, handling macOS Unicode filename issues.
    macOS screen recordings often have Unicode narrow no-break spaces (U+202F).
    """
    # Direct path works
    if os.path.isfile(video_arg):
        return os.path.abspath(video_arg)

    # Try glob matching in the same directory
    parent = os.path.dirname(video_arg) or "."
    basename = os.path.basename(video_arg)
    ext = os.path.splitext(basename)[1]

    # Try matching by extension in the directory
    if ext:
        matches = glob.glob(os.path.join(parent, f"*{ext}"))
        if len(matches) == 1:
            return os.path.abspath(matches[0])
        # Try fuzzy match
        name_start = basename[:10]
        for m in matches:
            if name_start in os.path.basename(m):
                return os.path.abspath(m)

    return None


def extract_frames(video_path: str, output_dir: str, fps: int = DEFAULT_FPS) -> int:
    """Extract frames from video using ffmpeg."""
    os.makedirs(output_dir, exist_ok=True)

    # Clean previous frames
    for f in glob.glob(os.path.join(output_dir, "frame_*.png")):
        os.remove(f)

    # Handle problematic filenames by creating a temp symlink
    symlink_path = None
    test_cmd = subprocess.run(
        ["ffprobe", "-v", "quiet", video_path],
        capture_output=True, timeout=10
    )
    if test_cmd.returncode != 0:
        # Create symlink to work around Unicode filename issues
        ext = os.path.splitext(video_path)[1]
        symlink_path = os.path.join(os.path.dirname(video_path), f"_temp_recording{ext}")
        try:
            os.symlink(video_path, symlink_path)
            video_path = symlink_path
        except OSError:
            pass

    cmd = [
        "ffmpeg", "-i", video_path,
        "-vf", f"fps={fps}",
        "-q:v", "2",
        os.path.join(output_dir, "frame_%04d.png"),
        "-y",  # overwrite
    ]

    print(f"  Extracting frames at {fps} fps...")
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)

    # Cleanup symlink
    if symlink_path and os.path.islink(symlink_path):
        os.remove(symlink_path)

    if result.returncode != 0:
        print(f"  ERROR: ffmpeg failed:\n{result.stderr[-500:]}")
        return 0

    frame_count = len(glob.glob(os.path.join(output_dir, "frame_*.png")))
    return frame_count


# ---------------------------------------------------------------------------
# LLM Integration
# ---------------------------------------------------------------------------

def encode_image_base64(image_path: str) -> str:
    """Encode image to base64 string."""
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def get_sampled_frames(frames_dir: str, interval: int = FRAME_SAMPLE_INTERVAL) -> list:
    """Get evenly sampled frame paths."""
    all_frames = sorted(glob.glob(os.path.join(frames_dir, "frame_*.png")))
    sampled = [all_frames[i] for i in range(0, len(all_frames), interval)]
    # Always include first and last
    if all_frames and all_frames[-1] not in sampled:
        sampled.append(all_frames[-1])
    return sampled


def analyze_with_anthropic(frames: list, principles: str, api_key: str) -> str:
    """Analyze frames using Anthropic Claude API."""
    try:
        import anthropic
    except ImportError:
        print("ERROR: anthropic package not installed. Run: pip install anthropic")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    # Build content blocks with images
    content = []
    content.append({
        "type": "text",
        "text": (
            "You are an expert QA engineer. I'm showing you screenshots extracted from a screen recording "
            "of someone using a web application. Analyze these frames and identify:\n"
            "1. The application name and URL\n"
            "2. Every distinct screen/page shown\n"
            "3. Every user action performed (clicks, form entries, selections, navigation)\n"
            "4. Every UI element interacted with (buttons, dropdowns, forms, modals, dialogs)\n"
            "5. Every state change (status badges, progress bars, success/error messages, toasts)\n"
            "6. The user roles involved (if visible)\n"
            "7. The complete user flow from start to finish\n\n"
            "The frames are in chronological order. Describe the complete flow in detail."
        ),
    })

    for i, frame_path in enumerate(frames):
        img_data = encode_image_base64(frame_path)
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": img_data,
            },
        })
        content.append({
            "type": "text",
            "text": f"[Frame {i+1} of {len(frames)}]",
        })

    print(f"  Sending {len(frames)} frames to Claude for analysis...")
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=8000,
        messages=[{"role": "user", "content": content}],
    )

    return response.content[0].text


def analyze_with_openai(frames: list, principles: str, api_key: str) -> str:
    """Analyze frames using OpenAI GPT-4o API."""
    try:
        import openai
    except ImportError:
        print("ERROR: openai package not installed. Run: pip install openai")
        sys.exit(1)

    client = openai.OpenAI(api_key=api_key)

    # Build content blocks
    content = []
    content.append({
        "type": "text",
        "text": (
            "You are an expert QA engineer. I'm showing you screenshots extracted from a screen recording "
            "of someone using a web application. Analyze these frames and identify:\n"
            "1. The application name and URL\n"
            "2. Every distinct screen/page shown\n"
            "3. Every user action performed (clicks, form entries, selections, navigation)\n"
            "4. Every UI element interacted with (buttons, dropdowns, forms, modals, dialogs)\n"
            "5. Every state change (status badges, progress bars, success/error messages, toasts)\n"
            "6. The user roles involved (if visible)\n"
            "7. The complete user flow from start to finish\n\n"
            "The frames are in chronological order. Describe the complete flow in detail."
        ),
    })

    for frame_path in frames:
        img_data = encode_image_base64(frame_path)
        content.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:image/png;base64,{img_data}",
                "detail": "high",
            },
        })

    print(f"  Sending {len(frames)} frames to GPT-4o for analysis...")
    response = client.chat.completions.create(
        model="gpt-4o",
        max_tokens=8000,
        messages=[{"role": "user", "content": content}],
    )

    return response.choices[0].message.content


def _pick_gemini_model() -> str:
    """
    Auto-select the best available Gemini model for this API key.
    Queries the live model list first; falls back to a priority list.
    """
    try:
        import google.generativeai as genai
        available = {m.name.split("/")[-1] for m in genai.list_models()
                     if "generateContent" in getattr(m, "supported_generation_methods", [])}
        priority = [
            "gemini-2.5-flash",       # Best free-tier price/performance
            "gemini-2.5-flash-lite",  # Fastest & cheapest
            "gemini-2.5-pro",         # Most capable (may need billing)
            "gemini-2.0-flash-lite",  # Stable fallback
            "gemini-2.0-flash-001",   # Pinned stable version
            "gemini-flash-latest",    # Latest flash alias
            "gemini-1.5-flash-latest",
            "gemini-1.5-flash",
        ]
        for name in priority:
            if name in available:
                print(f"  Auto-selected Gemini model: {name}")
                return name
        # Last resort: return first available vision-capable model
        if available:
            chosen = sorted(available)[0]
            print(f"  Using Gemini model: {chosen}")
            return chosen
    except Exception:
        pass

    # Hardcoded fallback if list_models() fails
    return "gemini-1.5-flash-latest"


def analyze_with_gemini(frames: list, principles: str, api_key: str) -> str:
    """Analyze frames using Google Gemini API."""
    try:
        import google.generativeai as genai
        from PIL import Image
    except ImportError:
        print("ERROR: Required packages not installed. Run: pip install google-generativeai pillow")
        sys.exit(1)

    genai.configure(api_key=api_key)

    model_name = _pick_gemini_model()
    model = genai.GenerativeModel(model_name)

    prompt = (
        "You are an expert QA engineer. I'm showing you screenshots extracted from a screen recording "
        "of someone using a web application. Analyze these frames and identify:\n"
        "1. The application name and URL\n"
        "2. Every distinct screen/page shown\n"
        "3. Every user action performed (clicks, form entries, selections, navigation)\n"
        "4. Every UI element interacted with (buttons, dropdowns, forms, modals, dialogs)\n"
        "5. Every state change (status badges, progress bars, success/error messages, toasts)\n"
        "6. The user roles involved (if visible)\n"
        "7. The complete user flow from start to finish\n\n"
        "The frames are in chronological order. Describe the complete flow in detail."
    )

    parts = [prompt]
    for frame_path in frames:
        parts.append(Image.open(frame_path))

    print(f"  Sending {len(frames)} frames to Gemini ({model_name}) for analysis...")

    # Retry with backoff on rate limit (429)
    for attempt in range(1, 4):
        try:
            response = model.generate_content(parts)
            return response.text
        except Exception as e:
            err = str(e)
            if "429" in err or "ResourceExhausted" in err or "quota" in err.lower():
                # Parse suggested retry delay from error message if available
                wait = 60 * attempt
                import re as _re
                m = _re.search(r"retry in (\d+)", err)
                if m:
                    wait = int(m.group(1)) + 5
                if attempt < 3:
                    print(f"  Rate limit hit. Waiting {wait}s before retry {attempt}/3...")
                    time.sleep(wait)
                else:
                    print("  ERROR: Gemini rate limit exceeded after 3 attempts.")
                    print("  The free tier allows ~2 requests/min. Try again in a few minutes.")
                    print("  Or enable billing at: https://console.cloud.google.com/billing")
                    sys.exit(1)
            else:
                raise


def generate_test_cases(flow_analysis: str, principles: str, provider: str, api_key: str) -> str:
    """Generate structured test cases from the flow analysis."""
    prompt = f"""You are an expert QA test case writer. Based on the application flow analysis below, generate comprehensive test cases.

## APPLICATION FLOW ANALYSIS:
{flow_analysis}

## TEST CASE WRITING PRINCIPLES:
{principles}

## INSTRUCTIONS:
Generate test cases in this EXACT format as a markdown table with these columns:
Test Case ID | Smoke | Sanity | Regression | E2E | Test Case | Test Case Automated | System | User Persona | Functional Area | Pre-condition | Expected Outcome/Response | Actual Outcome/Response

Rules:
- Write each test case title starting with "Verify that..." or "Verify if the user (role) can..."
- ONE test case = ONE specific verification
- Include the user role in parentheses: (upline), (downline), (admin), etc.
- Use ☑ for applicable test types, ☐ for not applicable
- Cover ALL of these scenarios:
  1. POSITIVE: Every happy path action observed in the flow
  2. NEGATIVE: Invalid inputs, empty fields, cancel actions, decline flows, missing data
  3. EDGE CASES: Boundary conditions, toggle behaviors, navigation mid-form, cross-entity isolation
- Group test cases by functional area with a clear ID prefix (e.g., TC-CC for Commission Config)
- Be granular: each button click, each form field, each dropdown, each dialog = separate test case
- Include pre-conditions specific to each test
- Expected outcomes should reference exact UI text, messages, and status changes observed

Generate at LEAST 80 test cases covering every interaction in the flow.

Output ONLY the markdown table (with header row and separator row), nothing else before or after."""

    if provider == "anthropic":
        try:
            import anthropic
        except ImportError:
            print("ERROR: anthropic package not installed.")
            sys.exit(1)

        client = anthropic.Anthropic(api_key=api_key)
        print("  Generating test cases with Claude...")
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=16000,
            messages=[{"role": "user", "content": prompt}],
        )
        return response.content[0].text

    elif provider == "openai":
        try:
            import openai
        except ImportError:
            print("ERROR: openai package not installed.")
            sys.exit(1)

        client = openai.OpenAI(api_key=api_key)
        print("  Generating test cases with GPT-4o...")
        response = client.chat.completions.create(
            model="gpt-4o",
            max_tokens=16000,
            messages=[{"role": "user", "content": prompt}],
        )
        return response.choices[0].message.content

    elif provider == "gemini":
        try:
            import google.generativeai as genai
        except ImportError:
            print("ERROR: google-generativeai not installed.")
            sys.exit(1)

        genai.configure(api_key=api_key)
        model_name = _pick_gemini_model()
        model = genai.GenerativeModel(model_name)
        print(f"  Generating test cases with Gemini ({model_name})...")
        import re as _re
        for attempt in range(1, 4):
            try:
                response = model.generate_content(prompt)
                return response.text
            except Exception as e:
                err = str(e)
                if "429" in err or "ResourceExhausted" in err or "quota" in err.lower():
                    wait = 60 * attempt
                    m = _re.search(r"retry in (\d+)", err)
                    if m:
                        wait = int(m.group(1)) + 5
                    if attempt < 3:
                        print(f"  Rate limit hit. Waiting {wait}s before retry {attempt}/3...")
                        time.sleep(wait)
                    else:
                        print("  ERROR: Gemini rate limit exceeded. Try again in a few minutes.")
                        sys.exit(1)
                else:
                    raise

    return ""


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def parse_markdown_table(md_text: str) -> list:
    """Parse markdown table rows into list of dicts."""
    rows = []
    headers = []
    for line in md_text.strip().split("\n"):
        line = line.strip()
        if not line.startswith("|"):
            continue
        cols = [c.strip() for c in line.split("|")[1:-1]]
        if not cols:
            continue
        # Skip separator row
        if all(set(c) <= {"-", ":", " "} for c in cols):
            continue
        if not headers:
            headers = cols
            continue
        if len(cols) >= len(headers):
            rows.append(dict(zip(headers, cols)))
    return headers, rows


def export_csv(headers: list, rows: list, output_path: str):
    """Export to CSV."""
    # Map checkbox symbols
    csv_headers = []
    for h in headers:
        csv_headers.append(h)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            csv_row = {}
            for h in headers:
                val = row.get(h, "")
                # Convert checkboxes to checkmarks
                if val == "☑":
                    val = "✓"
                elif val == "☐":
                    val = ""
                csv_row[h] = val
            writer.writerow(csv_row)
    print(f"  ✅ CSV: {output_path} ({len(rows)} test cases)")


def export_xlsx(headers: list, rows: list, output_path: str):
    """Export to styled Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  ⚠️  openpyxl not installed — skipping Excel. Run: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    hf = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    cf = Font(name="Arial", size=9)
    ca = Alignment(vertical="top", wrap_text=True)
    cca = Alignment(horizontal="center", vertical="top")
    fill1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    check_cols = {"Smoke", "Sanity", "Regression", "E2E", "Test Case Automated"}

    for ri, row in enumerate(rows, 2):
        rf = fill1 if ri % 2 == 0 else fill2
        for ci, h in enumerate(headers, 1):
            val = row.get(h, "")
            if val == "☑":
                val = "✓"
            elif val == "☐":
                val = ""
            c = ws.cell(row=ri, column=ci, value=val)
            c.font, c.border, c.fill = cf, tb, rf
            c.alignment = cca if h in check_cols else ca

    # Auto-size columns (rough estimate)
    for ci, h in enumerate(headers, 1):
        max_len = len(h)
        for row in rows[:5]:
            val = str(row.get(h, ""))
            max_len = max(max_len, min(len(val), 80))
        col_letter = chr(64 + ci) if ci <= 26 else f"{chr(64 + ci//26)}{chr(64 + ci%26)}"
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    ws.freeze_panes = "A2"
    last_col = chr(64 + len(headers)) if len(headers) <= 26 else "M"
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    wb.save(output_path)
    print(f"  ✅ Excel: {output_path} ({len(rows)} test cases)")


def export_markdown(headers: list, rows: list, output_path: str, video_name: str, flow_analysis: str):
    """Export to markdown with full structure."""
    with open(output_path, "w") as f:
        f.write(f"# Test Cases — Generated from Video\n\n")
        f.write(f"## Project Information\n")
        f.write(f"- **Date Created:** {time.strftime('%b %d, %Y')}\n")
        f.write(f"- **Source Video:** {video_name}\n")
        f.write(f"- **Generated By:** Clipcase\n\n")
        f.write(f"---\n\n")
        f.write(f"## Flow Analysis\n\n{flow_analysis}\n\n")
        f.write(f"---\n\n")
        f.write(f"## Test Cases\n\n")

        # Write table
        f.write("| " + " | ".join(headers) + " |\n")
        f.write("|" + "|".join(["---"] * len(headers)) + "|\n")
        for row in rows:
            vals = [row.get(h, "") for h in headers]
            f.write("| " + " | ".join(vals) + " |\n")

        f.write(f"\n---\n\n")
        f.write(f"## Summary\n\n")
        f.write(f"| Metric | Count |\n|--------|-------|\n")
        f.write(f"| Total Test Cases | {len(rows)} |\n")

        smoke = sum(1 for r in rows if r.get("Smoke") == "☑")
        sanity = sum(1 for r in rows if r.get("Sanity") == "☑")
        regression = sum(1 for r in rows if r.get("Regression") == "☑")
        e2e = sum(1 for r in rows if r.get("E2E") == "☑")
        f.write(f"| Smoke | {smoke} |\n")
        f.write(f"| Sanity | {sanity} |\n")
        f.write(f"| Regression | {regression} |\n")
        f.write(f"| E2E | {e2e} |\n")

    print(f"  ✅ Markdown: {output_path} ({len(rows)} test cases)")


# ---------------------------------------------------------------------------
# Main CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Clipcase — Convert screen recording videos into structured test cases.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python clipcase.py recording.mov
  python clipcase.py recording.mov --fps 2 --provider anthropic
  python clipcase.py recording.mov --fps 1 --provider openai --output my_tests
  python clipcase.py recording.mov --frames-only
        """,
    )
    parser.add_argument("video", help="Path to the video file (.mov, .mp4, .webm)")
    parser.add_argument("--fps", type=int, default=DEFAULT_FPS, help=f"Frames per second to extract (default: {DEFAULT_FPS})")
    parser.add_argument("--provider", choices=["anthropic", "openai", "gemini"], default="anthropic", help="LLM provider (default: anthropic)")
    parser.add_argument("--output", default=None, help="Output file base name (default: derived from video name)")
    parser.add_argument("--output-dir", default="output", help="Directory for generated test case files (default: output)")
    parser.add_argument("--frames-dir", default="frames", help="Directory to store extracted frames (default: frames)")
    parser.add_argument("--frames-only", action="store_true", help="Only extract frames, skip AI analysis")
    parser.add_argument("--sample-interval", type=int, default=FRAME_SAMPLE_INTERVAL, help=f"Sample every Nth frame for analysis (default: {FRAME_SAMPLE_INTERVAL})")
    parser.add_argument("--principles", default=None, help="Path to test case principles file (default: test_case_creation_principles.md)")

    args = parser.parse_args()

    # Load env vars
    load_env()

    print("=" * 60)
    print("  Clipcase — Video to Test Case Generator")
    print("=" * 60)

    # Step 1: Check ffmpeg
    print("\n[1/5] Checking ffmpeg...")
    if not check_ffmpeg():
        print("  ERROR: ffmpeg is not installed.")
        print("  Install it:")
        print("    macOS:   brew install ffmpeg")
        print("    Ubuntu:  sudo apt install ffmpeg")
        print("    Windows: choco install ffmpeg")
        sys.exit(1)
    print("  ✅ ffmpeg found")

    # Step 2: Resolve video file
    print("\n[2/5] Locating video file...")
    video_path = resolve_video_path(args.video)
    if not video_path:
        print(f"  ERROR: Video file not found: {args.video}")
        print("  Tip: macOS screen recordings may have hidden Unicode characters in filenames.")
        print("  Try using a glob pattern or renaming the file.")
        sys.exit(1)

    video_name = os.path.basename(video_path)
    print(f"  ✅ Found: {video_name}")

    # Get video info
    info = get_video_info(video_path)
    if info and "format" in info:
        duration = float(info["format"].get("duration", 0))
        mins, secs = divmod(int(duration), 60)
        print(f"  Duration: {mins}m {secs}s")
        for stream in info.get("streams", []):
            if stream.get("codec_type") == "video":
                print(f"  Resolution: {stream.get('width')}x{stream.get('height')}")
                break

    # Step 3: Extract frames
    print(f"\n[3/5] Extracting frames at {args.fps} fps...")
    frames_dir = os.path.join(os.path.dirname(video_path), args.frames_dir)
    frame_count = extract_frames(video_path, frames_dir, args.fps)
    if frame_count == 0:
        print("  ERROR: No frames extracted. Check video file.")
        sys.exit(1)
    print(f"  ✅ {frame_count} frames extracted to {frames_dir}/")

    if args.frames_only:
        print("\n  Done (--frames-only mode). Frames saved for manual review.")
        return

    # Step 4: Analyze frames with LLM
    print(f"\n[4/5] Analyzing frames with {args.provider}...")

    # Check API key
    if args.provider == "anthropic":
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            print("  ERROR: ANTHROPIC_API_KEY not set.")
            print("  Set it: export ANTHROPIC_API_KEY=sk-ant-...")
            print("  Or add to .env file: ANTHROPIC_API_KEY=sk-ant-...")
            sys.exit(1)
    elif args.provider == "openai":
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            print("  ERROR: OPENAI_API_KEY not set.")
            print("  Set it: export OPENAI_API_KEY=sk-...")
            print("  Or add to .env file: OPENAI_API_KEY=sk-...")
            sys.exit(1)
    elif args.provider == "gemini":
        api_key = os.environ.get("GEMINI_API_KEY")
        if not api_key:
            print("  ERROR: GEMINI_API_KEY not set.")
            print("  Set it: export GEMINI_API_KEY=AIza...")
            print("  Or add to .env file: GEMINI_API_KEY=AIza...")
            print("  Get a free key at: https://aistudio.google.com/app/apikey")
            sys.exit(1)

    # Load principles
    principles = ""
    principles_path = args.principles
    if not principles_path:
        # Look for default principles file
        default_path = os.path.join(os.path.dirname(video_path), "test_case_creation_principles.md")
        if os.path.exists(default_path):
            principles_path = default_path

    if principles_path and os.path.exists(principles_path):
        with open(principles_path) as f:
            principles = f.read()
        print(f"  Loaded principles: {os.path.basename(principles_path)}")
    else:
        print("  No principles file found (optional). Using default prompts.")

    # Sample frames
    sampled = get_sampled_frames(frames_dir, args.sample_interval)
    print(f"  Sampled {len(sampled)} frames (every {args.sample_interval}th frame)")

    # Send frames in batches for analysis
    all_analyses = []
    for batch_start in range(0, len(sampled), MAX_FRAMES_PER_BATCH):
        batch = sampled[batch_start:batch_start + MAX_FRAMES_PER_BATCH]
        batch_num = batch_start // MAX_FRAMES_PER_BATCH + 1
        total_batches = (len(sampled) + MAX_FRAMES_PER_BATCH - 1) // MAX_FRAMES_PER_BATCH
        print(f"  Analyzing batch {batch_num}/{total_batches} ({len(batch)} frames)...")

        if args.provider == "anthropic":
            analysis = analyze_with_anthropic(batch, principles, api_key)
        elif args.provider == "openai":
            analysis = analyze_with_openai(batch, principles, api_key)
        else:
            analysis = analyze_with_gemini(batch, principles, api_key)
        all_analyses.append(analysis)

    flow_analysis = "\n\n---\n\n".join(all_analyses)
    print(f"  ✅ Flow analysis complete")

    # Step 5: Generate test cases
    print(f"\n[5/5] Generating test cases...")
    test_case_md = generate_test_cases(flow_analysis, principles, args.provider, api_key)

    # Parse the generated table
    headers, rows = parse_markdown_table(test_case_md)
    if not rows:
        print("  WARNING: Could not parse test cases from LLM output.")
        print("  Saving raw output for manual review...")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        fallback_dir = os.path.join(script_dir, args.output_dir)
        os.makedirs(fallback_dir, exist_ok=True)
        raw_path = os.path.join(
            fallback_dir,
            (args.output or "test_cases") + "_raw.md",
        )
        with open(raw_path, "w") as f:
            f.write(test_case_md)
        print(f"  Saved raw output: {raw_path}")
        return

    print(f"  ✅ Generated {len(rows)} test cases")

    # Export
    base_name = args.output or os.path.splitext(video_name)[0].replace(" ", "_") + "_test_cases"
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, args.output_dir)
    os.makedirs(output_dir, exist_ok=True)

    md_path = os.path.join(output_dir, f"{base_name}.md")
    csv_path = os.path.join(output_dir, f"{base_name}.csv")
    xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")

    print(f"\n  Exporting...")
    export_markdown(headers, rows, md_path, video_name, flow_analysis)
    export_csv(headers, rows, csv_path)
    export_xlsx(headers, rows, xlsx_path)

    # Summary
    smoke = sum(1 for r in rows if r.get("Smoke") == "☑")
    regression = sum(1 for r in rows if r.get("Regression") == "☑")
    e2e = sum(1 for r in rows if r.get("E2E") == "☑")

    print(f"\n{'=' * 60}")
    print(f"  DONE — {len(rows)} test cases generated")
    print(f"  Smoke: {smoke} | Regression: {regression} | E2E: {e2e}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
